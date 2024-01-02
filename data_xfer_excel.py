#!/usr/bin/python3
""" This script movs data from an ESRI EXCEL spreadsheet to a MySql database
"""

import os
import argparse
import sys
from getpass import getpass
import openpyxl
from openpyxl import load_workbook
import mysql.connector

import a2database
from a2database import A2Database

# Check if we have the geometry transformation module
try:
    from osgeo import ogr
    from osgeo import osr
except ModuleNotFoundError:
    GEOM_CAN_TRANSFORM = False
else:
    GEOM_CAN_TRANSFORM = True

# The name of our script
SCRIPT_NAME = os.path.basename(__file__)

# Default host name to connect to the database
DEFAULT_HOST_NAME = 'localhost'

# Default number of heading lines in the EXCEL file
DEFAULT_NUM_HEADER_LINES = 1

# Default row that has the column names
DEFAULT_COL_NAMES_ROW = 1

# The default primary key database name for the sheet data
DEFAULT_PRIMARY_KEY_NAME = 'ObjectID'

# Default EPSG code for points
DEFAULT_GEOM_EPSG = 4326

# Argparse-related definitions
# Declare the progam description
ARGPARSE_PROGRAM_DESC = 'Uploads data from an ESRI generated excel spreadsheet to the ' \
                        'MySQL database'
# Epilog to argparse arguments
ARGPARSE_EPILOG = 'Duplicates are avoided by checking if the data already exists in the database'
# Host name help
ARGPARSE_HOST_HELP = f'The database host to connect to (default={DEFAULT_HOST_NAME})'
# Name of the database to connect to
ARGPARSE_DATABASE_HELP = 'The database to connect to'
# User name help
ARGPARSE_USER_HELP = 'The username to connect to the database with'
# Password help
ARGPARSE_PASSWORD_HELP = 'The password used to connect to the database (leave empty to be prompted)'
# Declare the help text for the EXCEL filename parameter (for argparse)
ARGPARSE_EXCEL_FILE_HELP = 'Path to the EXCEL file to upload'
# Declare the help text for the force deletion flag
ARGPARSE_UPDATE_HELP = 'Update existing data with new values (default is to skip updates)'
# Declare the help for no headers
ARGPARSE_HEADER_HELP = 'Specify the number of lines to consider as headings ' \
                       f'(default {DEFAULT_NUM_HEADER_LINES} lines)'
# Declare the help for where the column names are
ARGPARSE_COL_NAMES_ROW_HELP = 'Specify the row that contains the column names ' \
                              f'(default row is {DEFAULT_COL_NAMES_ROW})'
# Help text for specifying all the column names for the excel file
ARGPARSE_COL_NAMES_HELP = 'Comma seperated list of column names (used when column names are not ' \
                          'specified in the file)'
# Help text fo specifying the primary key to use
ARGPARSE_PRIMARY_KEY_HELP = 'The name of the primary key to use when checking if data is ' \
                            f'already in the database (default "{DEFAULT_PRIMARY_KEY_NAME}"). ' \
                            'Also needs to be a column in the spreadsheet'
# Help for specifying columns that are to be considered point columns
ARGPARSE_POINT_COLS_HELP = 'The names of the X and Y columns in the spreadsheet that represent ' \
                           'a point type when creating a schema ("<x name>,<y name>")'
# Help for specifying the EPSG code that the point coordinates are in
ARGPARSE_GEOMETRY_EPSG_HELP = 'The EPSG code of the coordinate system for the geometric values ' \
                           f'(default is {DEFAULT_GEOM_EPSG})'
# Help text for verbose flag
ARGPARSE_VERBOSE_HELP = 'Display additional information as the script is run'


def get_arguments() -> tuple:
    """ Returns the data from the parsed command line arguments
    Returns:
        A tuple consisting of a string containing the EXCEL file name to process, and
        a dict of the command line options
    Exceptions:
        A ValueError exception is raised if the filename is not specified
    Notes:
        If an error is found, the script will exit with a non-zero return code
    """
    parser = argparse.ArgumentParser(prog=SCRIPT_NAME,
                                     description=ARGPARSE_PROGRAM_DESC,
                                     epilog=ARGPARSE_EPILOG)
    parser.add_argument('excel_file', nargs='+',
                        help=ARGPARSE_EXCEL_FILE_HELP)
    parser.add_argument('-o', '--host', default=DEFAULT_HOST_NAME,
                        help=ARGPARSE_HOST_HELP)
    parser.add_argument('-d', '--database', help=ARGPARSE_DATABASE_HELP)
    parser.add_argument('-u', '--user', help=ARGPARSE_USER_HELP)
    parser.add_argument('-f', '--force', action='store_true',
                        help=ARGPARSE_UPDATE_HELP)
    parser.add_argument('--verbose', action='store_true',
                        help=ARGPARSE_VERBOSE_HELP)
    parser.add_argument('-p', '--password', action='store_true',
                        help=ARGPARSE_PASSWORD_HELP)
    parser.add_argument('--header', type=int, default=DEFAULT_NUM_HEADER_LINES,
                        help=ARGPARSE_HEADER_HELP)
    parser.add_argument('--col_names_row', type=int, default=DEFAULT_COL_NAMES_ROW,
                        help=ARGPARSE_COL_NAMES_ROW_HELP)
    parser.add_argument('--col_names', help=ARGPARSE_COL_NAMES_HELP)
    parser.add_argument('--point_cols', help=ARGPARSE_POINT_COLS_HELP)
    parser.add_argument('--geometry_epsg', type=int, default=DEFAULT_GEOM_EPSG,
                        help=ARGPARSE_GEOMETRY_EPSG_HELP)
    parser.add_argument('-k', '--key_name', default=DEFAULT_PRIMARY_KEY_NAME,
                        help=ARGPARSE_PRIMARY_KEY_HELP)
    args = parser.parse_args()

    # Find the EXCEL file and the password (which is allowed to be eliminated)
    excel_file = None
    if not args.excel_file:
        # Raise argument error
        raise ValueError('Missing a required argument')

    if len(args.excel_file) == 1:
        excel_file = args.excel_file[0]
    else:
        # Report the problem
        print('Too many arguments specified for input file', flush=True)
        parser.print_help()
        sys.exit(10)

    # Check that we can access the EXCEL file
    try:
        with open(excel_file, encoding="utf-8"):
            pass
    except FileNotFoundError:
        print(f'Unable to open EXCEL file {excel_file}', flush=True)
        sys.exit(11)

    print('HACK: ',args.point_cols)

    cmd_opts = {'force': args.force,
                'verbose': args.verbose,
                'host': args.host,
                'database': args.database,
                'user': args.user,
                'password': args.password,
                'header_lines': args.header,
                'col_names_row': args.col_names_row,
                'col_names': tuple(one_name.strip() for one_name in args.col_names.split(',')) \
                                        if args.col_names else None,
                'point_col_x': args.point_cols.split(',')[0] if args.point_cols else None,
                'point_col_y': args.point_cols.split(',')[1] if args.point_cols else None,
                'geometry_epsg': args.geometry_epsg,
                'primary_key': args.key_name
               }

    # Return the loaded JSON
    return excel_file, cmd_opts


def transform_points(from_epsg: int, to_epsg: int, values: tuple) -> list:
    """Returns the list with the last num_values converted to the new coordinate system
    Arguments:
        from_epsg: original EPSG code
        to_epsg: EPSG code to transform the points to
        values: the list of values to transform (assumes X1, Y1, X2, Y2, ...)
    Returns:
        The list of transformed values
    """
    if len(values) % 2 or len(values) < 1:
        raise ValueError('Unable to transform points, an even number of X,Y values pairs ' \
                         'are needed')

    # Make sure we can transform points
    if not GEOM_CAN_TRANSFORM:
        raise ValueError('Unable to transform points, supporting osgeo module is not installed')

    # Transform the points
    return_values = []
    idx = 0
    from_sr = osr.SpatialReference()
    from_sr.ImportFromEPSG(int(from_epsg))
    to_sr = osr.SpatialReference()
    to_sr.ImportFromEPSG(int(to_epsg))
    transform = osr.CreateCoordinateTransformation(from_sr, to_sr)
    while idx < len(values):
        cur_geom = ogr.CreateGeometryFromWkt(f'POINT({values[idx]} {values[idx+1]} {from_epsg})')
        cur_geom.Transform(transform)
        return_values.extend((cur_geom.GetX(), cur_geom.GetY()))
        idx = idx + 2

    return return_values


def transform_geom_cols(col_names: tuple, col_values: tuple, geom_col_info: dict, from_epsg: int, \
                        to_epsg: int) -> list:
    """Transforms geometry point to the specified coordinate system
    Arguments:
        col_name: the column names of the table
        col_values: the values associated with the column names
        geom_col_info: the geometry column information
        from_epsg: the EPSG code to tranform from
        to_epsg: the EPSG code to transform to
    Return:
        Returns a list of column values with the geometry values transformed
    """
    # Check if we can avoid the transformations
    if from_epsg == to_epsg:
        return list(col_values)
    if len(geom_col_info['sheet_cols']) % 2:
        raise ValueError('Invalid number of X,Y column name pairs specified')

    pt_indexes = []
    pt_values = []
    idx = 0
    while idx < len(geom_col_info['sheet_cols']):
        x_idx = col_names.index(geom_col_info['sheet_cols'][idx])
        y_idx = col_names.index(geom_col_info['sheet_cols'][idx+1])
        pt_indexes.append(x_idx)
        pt_indexes.append(y_idx)
        pt_values.append(col_values[x_idx])
        pt_values.append(col_values[y_idx])
        idx += 2

    new_pt_values = transform_points(from_epsg, to_epsg, pt_values)

    return_values = list(col_values)
    for idx, idx_val in enumerate(pt_indexes):
        return_values[idx_val] = new_pt_values[idx]

    return return_values


def process_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet, conn: A2Database, opts: dict) \
                    -> None:
    """Uploads the data in the worksheet
    Arguments:
        sheet: the worksheet to upload
        conn: the database connection
        opts: additional options
    """
    # Get the table name from the sheet title
    table_name = '_'.join(sheet.title.split('_')[:-1])

    print(f'Updating table {table_name} from tab {sheet.title}', flush=True)

    # Get the rows iterator
    rows_iter = sheet.iter_rows()

    # Get the column names
    if 'col_names' in opts and opts['col_names']:
        # User specified column names
        col_names = opts['col_names']
    else:
        # Find the row with the column names
        col_names = []
        # Skip to the row with the names
        cnt = 1
        while cnt < opts['col_names_row']:
            _ = next(rows_iter)
        for one_col in next(rows_iter):
            col_names.append(one_col.value)

    # Find geometry columns
    geom_col_info, col_alias = conn.get_col_info(table_name, col_names, opts['geometry_epsg'],
                                            colX1=opts['point_col_x'], rowY1=opts['point_col_y'])

    # Process the rows
    skipped_rows = 0
    added_updated_rows = 0
    for one_row in rows_iter:
        col_values = tuple(one_cell.value for one_cell in one_row)

        # Check for existing data and skip this row if it exists and we're not forcing
        data_exists = conn.check_data_exists(table_name, col_names, col_values,
                                            geom_col_info=geom_col_info,
                                            primary_key=opts['primary_key'],
                                            verbose=opts['verbose'] if 'verbose' in opts else False)
        if data_exists and not opts['force']:
            skipped_rows = skipped_rows + 1
            continue

        added_updated_rows = added_updated_rows + 1
        if geom_col_info and conn.epsg != opts['geometry_epsg']:
            col_values = transform_geom_cols(col_names, col_values, geom_col_info, \
                                             opts['geometry_epsg'], conn.epsg)
        conn.add_update_data(table_name, col_names, col_values, col_alias, geom_col_info, \
                             update=data_exists, \
                             primary_key=opts['primary_key'], verbose=opts['verbose'])

    if skipped_rows:
        print('    Processed', added_updated_rows + skipped_rows, \
                        f'rows with {skipped_rows} not updated', flush=True)
    else:
        print('    Processed', added_updated_rows + skipped_rows, 'rows', flush=True)


def load_excel_file(filepath: str, opts: dict) -> None:
    """Loads the excel file into the database
    Arguments:
        filepath: the path to the file to load
        opts: additional options
    """
    required_opts = ('host', 'database', 'password', 'user')

    # Check the opts parameter
    if opts is None:
        raise ValueError('Missing command line parameters')
    if not all(required in opts for required in required_opts):
        print('Missing required command line database parameters', flush=True)
        sys.exit(100)

    # Get the user password if they need to specify one
    if opts['password'] is not False:
        opts['password'] = getpass()

    # MySQL connection
    try:
        db_conn = a2database.connect(
            host=opts["host"],
            database=opts["database"],
            password=opts["password"],
            user=opts["user"]
        )
    except mysql.connector.errors.ProgrammingError as ex:
        print('Error', ex, flush=True)
        print('Please correct errors and try again', flush=True)
        sys.exit(101)

    # Set the default database EPSG
    db_conn.epsg = DEFAULT_GEOM_EPSG

    # Open the EXCEL file and process each tab
    workbook = load_workbook(filename=filepath, read_only=True, data_only=True)

    print(f'Updating using {filepath}')

    for one_sheet in workbook.worksheets:
        process_sheet(one_sheet, db_conn, opts)

    db_conn.commit()


if __name__ == '__main__':
    excel_filename, user_opts = get_arguments()
    load_excel_file(excel_filename, user_opts)
